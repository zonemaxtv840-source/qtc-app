# app.py - QTC Smart Sales Pro v4.5 (STOCK SEPARADO)
# Con soporte para XIAOMI, UGREEN y OTRAS marcas
# CORREGIDO: Stock SIN SUMAR - muestra cada almacén por separado
# CORREGIDO: Cards con texto negro
# CORREGIDO: UGREEN ahora lee stock REAL desde APRI.001 (columna "Disponible")
# CORREGIDO: Búsqueda inteligente UGREEN busca en stock si no encuentra en catálogo
# CORREGIDO: Badge "Sin stock" con texto en color negro

import streamlit as st
import pandas as pd
import re
import io
from datetime import datetime
import warnings
from typing import List, Dict, Optional, Tuple
from difflib import SequenceMatcher

warnings.filterwarnings('ignore')

# ============================================
# CONFIGURACIÓN DE PÁGINA
# ============================================

st.set_page_config(
    page_title="QTC Smart Sales Pro",
    page_icon="💼",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================
# CSS - TEXTO NEGRO EN CARDS
# ============================================

st.markdown("""
<style>
    /* FONDO DE PÁGINA */
    .stApp {
        background: linear-gradient(135deg, #0d47a1 0%, #1565c0 50%, #1e88e5 100%);
    }
    
    /* SIDEBAR */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #f8a35e 0%, #e87a2d 50%, #d45a1a 100%);
        border-right: 1px solid #ffcc80;
    }
    [data-testid="stSidebar"] * {
        color: #ffffff !important;
    }
    
    /* TEXTOS GENERALES */
    .stMarkdown, .stText, .stNumberInput label, .stSelectbox label {
        color: #ffffff !important;
    }
    
    h1, h2, h3 {
        color: #ffffff !important;
    }
    
    /* ========== CARDS - TEXTO NEGRO FORZADO ========== */
    div[style*="background:white"],
    div[style*="background:#ffffff"],
    div[style*="background:#fff"],
    div[style*="border-radius:16px"] {
        color: #1a1a2e !important;
    }
    
    div[style*="background:white"] *,
    div[style*="background:#ffffff"] *,
    div[style*="border-radius:16px"] * {
        color: #1a1a2e !important;
    }
    
    /* EXCEPCIÓN: Badges mantienen texto blanco */
    .badge-yessica, .badge-apri004, .badge-apri001, .badge-ugreen, .badge-warning,
    .badge-yessica *, .badge-apri004 *, .badge-apri001 *, .badge-ugreen *, .badge-warning * {
        color: white !important;
    }
    
    /* EXCEPCIÓN: Precios en naranja */
    .precio-texto {
        color: #e67e22 !important;
        font-weight: bold;
    }
    
    /* BADGES */
    .badge-yessica { background: #4CAF50; color: white; padding: 4px 12px; border-radius: 20px; font-size: 0.7rem; display: inline-block; margin: 2px; }
    .badge-apri004 { background: #FF9800; color: white; padding: 4px 12px; border-radius: 20px; font-size: 0.7rem; display: inline-block; margin: 2px; }
    .badge-apri001 { background: #f44336; color: white; padding: 4px 12px; border-radius: 20px; font-size: 0.7rem; display: inline-block; margin: 2px; }
    .badge-ugreen { background: #00BCD4; color: white; padding: 4px 12px; border-radius: 20px; font-size: 0.7rem; display: inline-block; margin: 2px; }
    .badge-warning { background: #f44336; color: white; padding: 4px 12px; border-radius: 20px; font-size: 0.7rem; display: inline-block; margin: 2px; }
    
    /* CONTADORES */
    .counter-summary {
        background: rgba(0,0,0,0.3);
        border-radius: 12px;
        padding: 1rem;
        margin-bottom: 1rem;
        display: flex;
        justify-content: space-around;
        flex-wrap: wrap;
    }
    .counter-item {
        text-align: center;
        padding: 0.5rem;
    }
    .counter-number {
        font-size: 1.5rem;
        font-weight: bold;
        margin-left: 8px;
    }
    
    /* FOOTER */
    .footer {
        text-align: center;
        padding: 1rem;
        color: rgba(255,255,255,0.7);
        font-size: 0.7rem;
        border-top: 1px solid rgba(255,255,255,0.2);
        margin-top: 2rem;
    }
</style>
""", unsafe_allow_html=True)

# ============================================
# FUNCIONES DE UTILIDAD
# ============================================

def corregir_numero(valor) -> float:
    if pd.isna(valor) or str(valor).strip() in ["", "0", "0.0", "-"]:
        return 0.0
    s = str(valor).upper().replace('S/', '').replace('$', '').replace(' ', '').strip()
    if ',' in s and '.' in s:
        s = s.replace(',', '')
    elif ',' in s:
        partes = s.split(',')
        if len(partes[-1]) <= 2:
            s = s.replace(',', '.')
        else:
            s = s.replace(',', '')
    s = re.sub(r'[^\d.]', '', s)
    try:
        return float(s)
    except:
        return 0.0

def limpiar_cabeceras(df: pd.DataFrame) -> pd.DataFrame:
    for i in range(min(20, len(df))):
        fila = [str(x).upper() for x in df.iloc[i].values]
        if any(h in item for h in ['SKU', 'COD', 'SAP', 'NUMERO', 'ARTICULO'] for item in fila):
            df.columns = [str(c).strip() for c in df.iloc[i]]
            return df.iloc[i+1:].reset_index(drop=True)
    return df

def cargar_archivo(uploaded_file) -> Optional[pd.DataFrame]:
    nombre = uploaded_file.name.lower()
    try:
        if nombre.endswith('.csv'):
            try:
                df = pd.read_csv(uploaded_file, encoding='utf-8')
            except:
                df = pd.read_csv(uploaded_file, encoding='latin-1')
        else:
            df = pd.read_excel(uploaded_file)
        return limpiar_cabeceras(df)
    except Exception as e:
        st.error(f"Error: {str(e)[:80]}")
        return None

def detectar_columna_sku(df: pd.DataFrame) -> str:
    posibles = ['SKU', 'COD', 'SAP', 'NUMERO', 'ARTICULO', 'CODIGO']
    for col in df.columns:
        col_upper = str(col).upper()
        for posible in posibles:
            if posible.upper() in col_upper:
                return col
    return df.columns[0]

def detectar_columna_descripcion(df: pd.DataFrame) -> str:
    posibles = ['DESC', 'DESCRIPCION', 'NOMBRE', 'PRODUCTO', 'GOODS']
    for col in df.columns:
        col_upper = str(col).upper()
        for posible in posibles:
            if posible.upper() in col_upper:
                return col
    return None

def detectar_columnas_precio(df: pd.DataFrame) -> Dict:
    precios = {}
    mapeo = {'P. IR': ['IR', 'MAYORISTA', 'MAYOR'], 
             'P. BOX': ['BOX', 'CAJA'], 
             'P. VIP': ['VIP']}
    
    for key, patrones in mapeo.items():
        for col in df.columns:
            col_upper = str(col).upper()
            for patron in patrones:
                if patron in col_upper:
                    precios[key] = col
                    break
            if key in precios:
                break
    
    if not precios and 'PRECIO' in [str(c).upper() for c in df.columns]:
        precios['P. VIP'] = 'PRECIO'
    
    return precios

def cargar_catalogo(archivo) -> Optional[Dict]:
    df = cargar_archivo(archivo)
    if df is None:
        return None
    
    return {
        'nombre': archivo.name,
        'df': df,
        'col_sku': detectar_columna_sku(df),
        'col_desc': detectar_columna_descripcion(df),
        'precios': detectar_columnas_precio(df)
    }

def cargar_stock(archivos, modo: str) -> List[Dict]:
    """
    Carga stock - SOLO usa la columna "Disponible" (o "Cantidad" como fallback)
    IGNORA: En stock, Comprometido, Solicitado
    """
    stocks = []
    
    for archivo in archivos:
        try:
            xls = pd.ExcelFile(archivo)
            for hoja in xls.sheet_names:
                hoja_upper = hoja.upper()
                if modo == "XIAOMI":
                    if not any(h in hoja_upper for h in ['APRI', 'YESSICA']):
                        continue
                else:
                    if 'APRI.001' not in hoja_upper:
                        continue
                
                df = pd.read_excel(archivo, sheet_name=hoja)
                df = limpiar_cabeceras(df)
                
                col_sku = detectar_columna_sku(df)
                col_cant = None
                
                # 1. Buscar columna "Disponible" (PRIORIDAD MÁXIMA)
                for col in df.columns:
                    if 'DISPONIBLE' in str(col).upper():
                        col_cant = col
                        break
                
                # 2. Si no hay "Disponible", buscar "Cantidad" (FALLBACK)
                if not col_cant:
                    for col in df.columns:
                        if 'CANTIDAD' in str(col).upper() or 'CANT' in str(col).upper():
                            col_cant = col
                            break
                
                # 3. Si no hay ninguna, error
                if not col_cant:
                    st.error(f"❌ Hoja {hoja}: No se encontró columna 'Disponible' ni 'Cantidad'")
                    continue
                
                stocks.append({
                    'nombre': f"{archivo.name} [{hoja}]",
                    'df': df,
                    'col_sku': col_sku,
                    'col_cant': col_cant,
                    'hoja': hoja
                })
                st.success(f"✅ {archivo.name} → {hoja} (usando: {col_cant})")
                
        except Exception as e:
            st.error(f"Error en {archivo.name}: {str(e)[:80]}")
    
    return stocks

# ============================================
# FUNCIONES DE SIMILITUD Y BÚSQUEDA
# ============================================

def normalizar_texto(texto: str) -> str:
    """Normaliza texto para mejor comparación"""
    if not texto:
        return ""
    texto = texto.lower().strip()
    
    correcciones = {
        "xioami": "xiaomi", "xiomi": "xiaomi", "xiamoi": "xiaomi",
        "earphone": "earphone", "earphones": "earphone",
    }
    for mal, bien in correcciones.items():
        texto = texto.replace(mal, bien)
    
    sufijos = [' - rn', ' - es', ' - us', ' - eu', ' - gl', ' - demo', ' - rr']
    for sufijo in sufijos:
        texto = texto.replace(sufijo, '')
    
    return texto.strip()

def calcular_similitud(texto1: str, texto2: str) -> float:
    if not texto1 or not texto2:
        return 0.0
    
    texto1 = normalizar_texto(texto1)
    texto2 = normalizar_texto(texto2)
    
    if texto1 == texto2:
        return 100.0
    
    palabras1 = set(texto1.split())
    palabras2 = set(texto2.split())
    
    interseccion = len(palabras1.intersection(palabras2))
    union = len(palabras1.union(palabras2))
    
    if union == 0:
        return 0.0
    
    jaccard = interseccion / union
    sequence_match = SequenceMatcher(None, texto1, texto2).ratio()
    
    similitud = (jaccard * 0.6 + sequence_match * 0.4) * 100
    return round(similitud, 1)

def buscar_stock_para_sku(sku: str, stocks: List[Dict]) -> Dict:
    """
    Busca stock - SOLO usa la columna "Disponible" o "Cantidad"
    CADA ALMACÉN POR SEPARADO (NO SUMA)
    """
    sku_limpio = sku.strip().upper()
    stock_yessica = 0
    stock_apri004 = 0
    stock_apri001 = 0
    
    for stock in stocks:
        df = stock['df']
        df_sku = df[stock['col_sku']].astype(str).str.strip().str.upper()
        mask = df_sku == sku_limpio
        if mask.any():
            row = df[mask].iloc[0]
            col_cant = stock.get('col_cant')
            
            if col_cant:
                # Leer SOLO el valor de la columna Disponible/Cantidad
                cantidad = int(corregir_numero(row[col_cant]))
                hoja = stock['hoja'].upper()
                
                # ASIGNAR, NO SUMAR
                if 'YESSICA' in hoja:
                    stock_yessica = cantidad
                elif 'APRI.004' in hoja:
                    stock_apri004 = cantidad
                elif 'APRI.001' in hoja:
                    stock_apri001 = cantidad
    
    return {
        'yessica': stock_yessica,
        'apri004': stock_apri004,
        'apri001': stock_apri001,
        'total': stock_yessica + stock_apri004 + stock_apri001
    }

def buscar_sku_por_descripcion(descripcion: str, catalogos: List[Dict], precio_key: str, umbral: float = 70.0) -> Optional[Dict]:
    """Busca un SKU en catálogos por descripción similar"""
    if not descripcion or not catalogos:
        return None
    
    desc_norm = normalizar_texto(descripcion)
    mejores_matches = []
    
    for cat in catalogos:
        df = cat['df']
        col_desc = cat.get('col_desc')
        if not col_desc:
            continue
        
        if precio_key not in cat.get('precios', {}):
            continue
        
        col_precio = cat['precios'][precio_key]
        
        for _, row in df.iterrows():
            desc_cat = normalizar_texto(str(row[col_desc]))
            similitud = calcular_similitud(desc_norm, desc_cat)
            
            if similitud >= umbral:
                try:
                    precio = float(row[col_precio]) if pd.notna(row[col_precio]) else 0
                    if precio > 0:
                        mejores_matches.append({
                            'precio': precio,
                            'sku_match': str(row[cat['col_sku']]).strip(),
                            'similitud': similitud,
                            'catalogo': cat['nombre'][:30]
                        })
                except:
                    pass
    
    if mejores_matches:
        mejores_matches.sort(key=lambda x: x['similitud'], reverse=True)
        return mejores_matches[0]
    
    return None

def buscar_producto(sku: str, catalogos: List[Dict], stocks: List[Dict], precio_key: str) -> Dict:
    """Busca producto - stock separado por almacén"""
    sku_limpio = sku.strip().upper()
    
    # PASO 1: BUSCAR STOCK (CADA ALMACÉN POR SEPARADO)
    stock_info = buscar_stock_para_sku(sku_limpio, stocks)
    
    # PASO 2: BUSCAR DESCRIPCIÓN Y PRECIO
    descripcion = f"SKU: {sku}"
    precio = 0.0
    sku_equivalente = None
    similitud_equivalente = 0
    precio_equivalente = 0
    
    for cat in catalogos:
        df = cat['df']
        df_sku = df[cat['col_sku']].astype(str).str.strip().str.upper()
        mask = df_sku == sku_limpio
        if mask.any():
            row = df[mask].iloc[0]
            if precio_key in cat['precios']:
                col_precio = cat['precios'][precio_key]
                precio = corregir_numero(row[col_precio])
            if cat['col_desc']:
                descripcion = str(row[cat['col_desc']])[:200]
            break
    
    stock_total = stock_info['yessica'] + stock_info['apri004'] + stock_info['apri001']
    
    # Si no encontró descripción pero tiene stock, buscarla en STOCK
    if descripcion == f"SKU: {sku}" and stock_total > 0:
        for stock in stocks:
            df = stock['df']
            df_sku = df[stock['col_sku']].astype(str).str.strip().str.upper()
            mask = df_sku == sku_limpio
            if mask.any():
                row = df[mask].iloc[0]
                for col in df.columns:
                    col_upper = str(col).upper()
                    if any(p in col_upper for p in ['DESC', 'DESCRIPCION', 'PRODUCTO', 'NOMBRE']):
                        desc_stock = str(row[col])[:200]
                        if desc_stock and desc_stock != 'nan':
                            descripcion = desc_stock
                            break
                break
    
    # PASO 3: SI TIENE STOCK PERO NO PRECIO → BUSCAR POR DESCRIPCIÓN
    if precio == 0 and stock_total > 0 and descripcion and descripcion != f"SKU: {sku}":
        match = buscar_sku_por_descripcion(descripcion, catalogos, precio_key, umbral=70.0)
        
        if match and match['precio'] > 0:
            sku_equivalente = match['sku_match']
            similitud_equivalente = match['similitud']
            precio_equivalente = match['precio']
    
    return {
        'sku': sku,
        'descripcion': descripcion,
        'precio': precio,
        'precio_equivalente': precio_equivalente,
        'stock_yessica': stock_info['yessica'],
        'stock_apri004': stock_info['apri004'],
        'stock_apri001': stock_info['apri001'],
        'stock_total': stock_total,
        'tiene_stock': stock_total > 0,
        'tiene_precio': precio > 0,
        'sku_equivalente': sku_equivalente,
        'similitud_equivalente': similitud_equivalente,
        'alternativas': []
    }

def construir_badge_stock(stock_yessica, stock_apri004, stock_apri001):
    """Construye badges - MUESTRA LOS TRES ALMACENES"""
    return f"""
    <div style="display:flex; flex-wrap:wrap; gap:8px; margin:8px 0;">
        <span class="badge-yessica">🟢 YESSICA: {stock_yessica}</span>
        <span class="badge-apri004">🟡 APRI.004: {stock_apri004}</span>
        <span class="badge-apri001">🔴 APRI.001: {stock_apri001}</span>
    </div>
    """

def generar_excel(items: List[Dict], cliente: str, ruc: str) -> bytes:
    output = io.BytesIO()
    
    df = pd.DataFrame(items)
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Cotizacion', index=False, startrow=6)
        
        workbook = writer.book
        ws = writer.sheets['Cotizacion']
        
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="e67e22", end_color="e67e22", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws['A1'] = 'QTC SMART SALES PRO'
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A3'] = 'FECHA:'
        ws['B3'] = datetime.now().strftime("%d/%m/%Y %H:%M")
        ws['A4'] = 'CLIENTE:'
        ws['B4'] = cliente.upper()
        ws['A5'] = 'RUC:'
        ws['B5'] = ruc
        
        headers = ['SKU', 'DESCRIPCIÓN', 'CANTIDAD', 'PRECIO UNIT.', 'TOTAL']
        for i, header in enumerate(headers, start=1):
            cell = ws.cell(row=7, column=i, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        
        for row_idx, item in enumerate(items, start=8):
            ws.cell(row=row_idx, column=1, value=item['sku']).border = border
            ws.cell(row=row_idx, column=2, value=item['descripcion']).border = border
            ws.cell(row=row_idx, column=3, value=item['cantidad']).border = border
            
            precio_cell = ws.cell(row=row_idx, column=4, value=item['precio'])
            precio_cell.number_format = '"S/." #,##0.00'
            precio_cell.border = border
            
            total_cell = ws.cell(row=row_idx, column=5, value=item['total'])
            total_cell.number_format = '"S/." #,##0.00'
            total_cell.border = border
        
        total_row = len(items) + 8
        total_label = ws.cell(row=total_row, column=4, value='TOTAL S/.')
        total_label.font = Font(bold=True, color="FFFFFF")
        total_label.fill = PatternFill(start_color="e67e22", end_color="e67e22", fill_type="solid")
        total_label.border = border
        
        total_valor = ws.cell(row=total_row, column=5, value=sum(item['total'] for item in items))
        total_valor.number_format = '"S/." #,##0.00'
        total_valor.border = border
        
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 110
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        
        ws.freeze_panes = 'A8'
    
    return output.getvalue()

# ============================================
# CARGAR UGREEN (MODIFICADO: SIN STOCK DEL CATÁLOGO)
# ============================================

def cargar_ugreen_catalogo(archivo) -> Optional[Dict]:
    df = cargar_archivo(archivo)
    if df is None:
        return None
    
    col_sku = None
    col_desc = None
    col_mayor = None
    col_caja = None
    col_vip = None
    
    for col in df.columns:
        col_upper = str(col).upper()
        if 'SKU' in col_upper:
            col_sku = col
        elif 'DESCRITPION' in col_upper or 'DESCRIPCION' in col_upper:
            col_desc = col
        elif col_upper == 'MAYOR':
            col_mayor = col
        elif col_upper == 'CAJA':
            col_caja = col
        elif col_upper == 'VIP':
            col_vip = col
    
    if not col_sku:
        col_sku = df.columns[0]
    
    precios = {}
    if col_mayor:
        precios['P. IR'] = col_mayor
    if col_caja:
        precios['P. BOX'] = col_caja
    if col_vip:
        precios['P. VIP'] = col_vip
    
    return {
        'nombre': archivo.name,
        'df': df,
        'col_sku': col_sku,
        'col_desc': col_desc,
        'precios': precios,
        'tipo': 'UGREEN'
    }


def buscar_stock_ugreen(sku: str, stocks: List[Dict]) -> int:
    """
    Busca stock REAL de UGREEN en APRI.001, columna "Disponible"
    """
    sku_limpio = sku.strip().upper()
    
    if not stocks:
        return 0
    
    for stock in stocks:
        hoja = stock.get('hoja', '').upper()
        nombre = stock.get('nombre', '').upper()
        
        if 'APRI.001' not in hoja and 'APRI.001' not in nombre:
            continue
        
        df = stock['df']
        col_sku = stock.get('col_sku')
        if not col_sku:
            col_sku = detectar_columna_sku(df)
        
        df_sku = df[col_sku].astype(str).str.strip().str.upper()
        mask = df_sku == sku_limpio
        
        if mask.any():
            row = df[mask].iloc[0]
            for col in df.columns:
                if 'DISPONIBLE' in str(col).upper():
                    return int(corregir_numero(row[col]))
            return 0
    
    return 0


def buscar_ugreen_producto(busqueda: str, ugreen_catalogo: Dict, stocks: List[Dict] = None) -> Optional[List[Dict]]:
    """
    Busca producto UGREEN:
    1. Primero en catálogo (por SKU o descripción)
    2. Si no encuentra, busca SOLO en stocks (APRI.001) por SKU exacto
    """
    resultados = []
    skus_encontrados = set()
    
    # ========== PASO 1: BUSCAR EN CATÁLOGO ==========
    if ugreen_catalogo:
        df = ugreen_catalogo['df']
        col_sku = ugreen_catalogo['col_sku']
        col_desc = ugreen_catalogo['col_desc']
        
        mask_sku = df[col_sku].astype(str).str.contains(busqueda, case=False, na=False)
        mask_desc = pd.Series([False] * len(df))
        if col_desc:
            mask_desc = df[col_desc].astype(str).str.contains(busqueda, case=False, na=False)
        
        mask = mask_sku | mask_desc
        coincidencias = df[mask]
        
        for _, row in coincidencias.iterrows():
            sku = str(row[col_sku]).strip().upper()
            skus_encontrados.add(sku)
            descripcion = str(row[col_desc])[:200] if col_desc else f"SKU: {sku}"
            
            precio_mayor = corregir_numero(row.get('Mayor', 0))
            precio_caja = corregir_numero(row.get('Caja', 0))
            precio_vip = corregir_numero(row.get('Vip', 0))
            
            stock = 0
            if stocks:
                stock = buscar_stock_ugreen(sku, stocks)
            
            resultados.append({
                'sku': sku,
                'descripcion': descripcion,
                'precios': {
                    'P. IR': precio_mayor,
                    'P. BOX': precio_caja,
                    'P. VIP': precio_vip,
                },
                'stock': stock,
                'tiene_stock': stock > 0,
                'tiene_precio': precio_vip > 0 or precio_caja > 0 or precio_mayor > 0,
                'tipo': 'UGREEN',
                'desde': 'catalogo'
            })
    
    # ========== PASO 2: BUSCAR SOLO EN STOCK SI NO ENCONTRÓ EN CATÁLOGO ==========
    if not resultados and stocks:
        # Buscar SKU exacto en stocks (APRI.001)
        sku_limpio = busqueda.strip().upper()
        
        for stock in stocks:
            hoja = stock.get('hoja', '').upper()
            nombre = stock.get('nombre', '').upper()
            
            if 'APRI.001' not in hoja and 'APRI.001' not in nombre:
                continue
            
            df = stock['df']
            col_sku = stock.get('col_sku')
            if not col_sku:
                col_sku = detectar_columna_sku(df)
            
            df_sku = df[col_sku].astype(str).str.strip().str.upper()
            mask = df_sku == sku_limpio
            
            if mask.any():
                row = df[mask].iloc[0]
                
                # Obtener descripción desde stock
                descripcion = f"SKU: {sku_limpio}"
                for col in df.columns:
                    col_upper = str(col).upper()
                    if any(p in col_upper for p in ['DESC', 'DESCRIPCION', 'PRODUCTO', 'NOMBRE']):
                        desc_stock = str(row[col])[:200]
                        if desc_stock and desc_stock != 'nan':
                            descripcion = desc_stock
                            break
                
                # Obtener stock
                stock_cantidad = 0
                for col in df.columns:
                    if 'DISPONIBLE' in str(col).upper():
                        stock_cantidad = int(corregir_numero(row[col]))
                        break
                
                resultados.append({
                    'sku': sku_limpio,
                    'descripcion': descripcion,
                    'precios': {'P. IR': 0, 'P. BOX': 0, 'P. VIP': 0},
                    'stock': stock_cantidad,
                    'tiene_stock': stock_cantidad > 0,
                    'tiene_precio': False,
                    'tipo': 'UGREEN',
                    'desde': 'stock_solo'
                })
                break  # Solo tomar el primer encuentro
    
    return resultados if resultados else None

# ============================================
# INICIALIZACIÓN DE SESIÓN
# ============================================

if 'auth' not in st.session_state:
    st.session_state.auth = False
if 'modo' not in st.session_state:
    st.session_state.modo = "XIAOMI"
if 'precio_key' not in st.session_state:
    st.session_state.precio_key = "P. VIP"
if 'catalogos' not in st.session_state:
    st.session_state.catalogos = []
if 'stocks' not in st.session_state:
    st.session_state.stocks = []
if 'carrito' not in st.session_state:
    st.session_state.carrito = []
if 'ugreen_catalogo' not in st.session_state:
    st.session_state.ugreen_catalogo = None

# ============================================
# LOGIN
# ============================================

if not st.session_state.auth:
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown('<div style="background:rgba(255,255,255,0.95);border-radius:20px;padding:2rem;">', unsafe_allow_html=True)
        
        try:
            st.image("logo.png", width=100)
        except:
            st.markdown("<h1 style='color:#e94560;text-align:center;'>QTC</h1>", unsafe_allow_html=True)
        
        st.markdown("<h2 style='color:#1a1a2e;text-align:center;'>QTC Smart Sales Pro</h2>", unsafe_allow_html=True)
        st.markdown("<p style='color:#666;text-align:center;'>Sistema Profesional de Cotización</p>", unsafe_allow_html=True)
        
        user = st.text_input("👤 Usuario", placeholder="admin / kimberly / vendedor")
        pw = st.text_input("🔒 Contraseña", type="password")
        
        col_btn1, col_btn2 = st.columns(2)
        with col_btn1:
            if st.button("🚀 Ingresar", use_container_width=True):
                credenciales = {
                    "admin": {"password": "qtc2026", "role": "ADMIN", "name": "Administrador"},
                    "kimberly": {"password": "kam2026", "role": "KAM", "name": "Kimberly"},
                    "vendedor": {"password": "ventas2026", "role": "VENDEDOR", "name": "Vendedor"}
                }
                if user in credenciales and pw == credenciales[user]["password"]:
                    st.session_state.auth = True
                    st.session_state.user_role = credenciales[user]["role"]
                    st.session_state.user_name = credenciales[user]["name"]
                    st.rerun()
                else:
                    st.error("❌ Credenciales incorrectas")
        with col_btn2:
            if st.button("👤 Invitado", use_container_width=True):
                st.session_state.auth = True
                st.session_state.user_role = "INVITADO"
                st.session_state.user_name = "Invitado"
                st.rerun()
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    st.stop()

# ============================================
# HEADER
# ============================================

col1, col2, col3 = st.columns([1, 5, 2])
with col1:
    try:
        st.image("logo.png", width=60)
    except:
        st.markdown("**QTC**", unsafe_allow_html=True)
with col2:
    st.markdown("# QTC Smart Sales Pro")
    st.caption("Sistema Profesional de Cotización | Soporte XIAOMI · UGREEN · OTRAS MARCAS")
with col3:
    role_badge = {"ADMIN": "🔧", "KAM": "⭐", "VENDEDOR": "🛒", "INVITADO": "👤"}
    badge = role_badge.get(st.session_state.user_role, "👤")
    st.markdown(f"""
    <div style="background: rgba(255,255,255,0.1); padding: 0.5rem 1rem; border-radius: 12px; text-align: right;">
        <span>{badge} {st.session_state.user_name}</span><br>
        <span style="font-size: 0.7rem;">{st.session_state.user_role}</span>
    </div>
    """, unsafe_allow_html=True)
    if st.button("🚪 Cerrar Sesión", key="logout"):
        st.session_state.auth = False
        st.session_state.carrito = []
        st.rerun()

st.markdown("---")

# ============================================
# SIDEBAR
# ============================================

with st.sidebar:
    st.markdown("### 🎯 Configuración")
    
    marca_seleccionada = st.radio(
        "📌 Marca / Modo",
        ["XIAOMI", "UGREEN", "OTRAS MARCAS"],
        index=0 if st.session_state.modo == "XIAOMI" else 1,
        help="XIAOMI: Stock YESSICA/APRI.004/APRI.001\nUGREEN: Catálogo específico"
    )
    st.session_state.modo = marca_seleccionada
    
    st.markdown("---")
    
    precio_opcion = st.radio(
        "💰 Nivel de precio",
        ["P. VIP", "P. BOX", "P. IR"],
        index=0
    )
    st.session_state.precio_key = precio_opcion
    
    st.markdown("---")
    
    st.markdown("### 📂 Archivos")
    
    if marca_seleccionada != "UGREEN":
        st.markdown("**📚 Catálogos de precios**")
        archivos_cat = st.file_uploader(
            "Excel o CSV",
            type=['xlsx', 'xls', 'csv'],
            accept_multiple_files=True,
            key="cat_upload"
        )
        if archivos_cat:
            st.session_state.catalogos = []
            for archivo in archivos_cat:
                cat = cargar_catalogo(archivo)
                if cat:
                    st.session_state.catalogos.append(cat)
                    st.success(f"✅ {archivo.name[:30]}")
    
    if marca_seleccionada == "UGREEN":
        st.markdown("**📚 Catálogo UGREEN**")
        archivo_ugreen = st.file_uploader(
            "Excel UGREEN",
            type=['xlsx', 'xls'],
            accept_multiple_files=False,
            key="ugreen_upload"
        )
        if archivo_ugreen:
            ugreen_cat = cargar_ugreen_catalogo(archivo_ugreen)
            if ugreen_cat:
                st.session_state.ugreen_catalogo = ugreen_cat
                st.success(f"✅ UGREEN: {archivo_ugreen.name[:30]}")
    
    st.markdown("**📦 Reportes de stock**")
    st.caption("📌 APRI.001 usa columna 'Disponible'")
    archivos_stock = st.file_uploader(
        "Excel",
        type=['xlsx', 'xls'],
        accept_multiple_files=True,
        key="stock_upload"
    )
    if archivos_stock:
        st.session_state.stocks = cargar_stock(archivos_stock, st.session_state.modo)
    
    st.markdown("---")
    
    if st.session_state.carrito:
        st.markdown(f"### 🛒 Carrito")
        st.metric("Productos", len(st.session_state.carrito))
        total = sum(item['total'] for item in st.session_state.carrito)
        st.metric("Total", f"S/ {total:,.2f}")
        
        if st.button("🧹 Limpiar carrito", use_container_width=True):
            st.session_state.carrito = []
            st.rerun()

# ============================================
# TABS PRINCIPALES
# ============================================

tab1, tab2, tab3, tab4 = st.tabs(["📦 MODO MASIVO (Bulk)", "🔍 BÚSQUEDA INTELIGENTE", "🛒 CARRITO DE COTIZACIÓN", "🔎 SMART CODE SEARCH"])

# ========== TAB 1: MODO MASIVO ==========
with tab1:
    st.markdown("### 📦 Ingresa productos en formato masivo")
    st.caption(f"🔍 Modo: **{st.session_state.modo}** | Formato: `SKU:CANTIDAD` (uno por línea)")
    
    texto_bulk = st.text_area(
        "",
        height=200,
        placeholder="Ejemplo:\nRN9401276NA8:100\nCN0200047BK8:50\nRN0200065BK8:25"
    )
    
    col_b1, col_b2 = st.columns([1, 1])
    with col_b1:
        if st.button("🚀 Procesar lista", type="primary", use_container_width=True):
            if not texto_bulk:
                st.warning("Ingresa productos")
            elif st.session_state.modo == "XIAOMI" and not st.session_state.catalogos:
                st.warning("Carga catálogos primero")
            elif st.session_state.modo == "XIAOMI" and not st.session_state.stocks:
                st.warning("Carga stock primero")
            else:
                if st.session_state.modo == "XIAOMI":
                    pedidos = []
                    for line in texto_bulk.strip().split('\n'):
                        line = line.strip()
                        if ':' in line:
                            parts = line.split(':')
                            if len(parts) == 2:
                                try:
                                    sku = parts[0].strip().upper()
                                    cant = int(parts[1].strip())
                                    if cant > 0:
                                        pedidos.append({'sku': sku, 'cantidad': cant})
                                except:
                                    pass
                    
                    if pedidos:
                        with st.spinner("Procesando..."):
                            resultados_procesados = []
                            encontrados = 0
                            con_precio = 0
                            con_stock = 0
                            
                            for pedido in pedidos:
                                prod = buscar_producto(pedido['sku'], st.session_state.catalogos, st.session_state.stocks, st.session_state.precio_key)
                                
                                if prod['tiene_precio']:
                                    encontrados += 1
                                    con_precio += 1
                                
                                if prod['tiene_stock']:
                                    con_stock += 1
                                
                                if prod['tiene_precio'] and prod['tiene_stock']:
                                    cantidad_cotizar = min(pedido['cantidad'], prod['stock_total'])
                                    estado = "✅ OK"
                                    if cantidad_cotizar < pedido['cantidad']:
                                        estado = f"⚠️ Stock insuficiente ({cantidad_cotizar}/{pedido['cantidad']})"
                                elif not prod['tiene_precio'] and prod['tiene_stock']:
                                    cantidad_cotizar = 0
                                    estado = "⚠️ Stock sin precio - Verificar SKU"
                                elif not prod['tiene_precio']:
                                    cantidad_cotizar = 0
                                    estado = "❌ Sin precio"
                                else:
                                    cantidad_cotizar = 0
                                    estado = "❌ Sin stock"
                                
                                resultados_procesados.append({
                                    **prod,
                                    'cantidad_solicitada': pedido['cantidad'],
                                    'cantidad_cotizar': cantidad_cotizar,
                                    'estado': estado
                                })
                            
                            st.session_state.resultados_bulk = resultados_procesados
                            
                            total_ingresados = len(pedidos)
                            total_encontrados = encontrados
                            total_con_stock = con_stock
                            total_sin_precio = total_ingresados - total_encontrados
                            total_sin_stock = total_ingresados - total_con_stock
                            
                            st.markdown(f"""
                            <div class="counter-summary">
                                <div class="counter-item">📋 Ingresados: <strong>{total_ingresados}</strong></div>
                                <div class="counter-item" style="background:#4CAF50;color:white;">✅ Con precio: <strong>{total_encontrados}</strong></div>
                                <div class="counter-item">📦 Con stock: <strong>{total_con_stock}</strong></div>
                                <div class="counter-item" style="color:#f44336;">❌ Sin precio: <strong>{total_sin_precio}</strong></div>
                                <div class="counter-item" style="color:#f44336;">🚫 Sin stock: <strong>{total_sin_stock}</strong></div>
                            </div>
                            """, unsafe_allow_html=True)
                            
                            st.success(f"✅ Procesados {len(pedidos)} productos")
                    else:
                        st.warning("No se encontraron productos válidos")
                
                elif st.session_state.modo == "UGREEN" and st.session_state.ugreen_catalogo:
                    pedidos = []
                    for line in texto_bulk.strip().split('\n'):
                        line = line.strip()
                        if ':' in line:
                            parts = line.split(':')
                            if len(parts) == 2:
                                try:
                                    sku = parts[0].strip().upper()
                                    cant = int(parts[1].strip())
                                    if cant > 0:
                                        pedidos.append({'sku': sku, 'cantidad': cant})
                                except:
                                    pass
                    
                    if pedidos:
                        with st.spinner("Procesando UGREEN..."):
                            resultados_procesados = []
                            for pedido in pedidos:
                                resultados_ugreen = buscar_ugreen_producto(pedido['sku'], st.session_state.ugreen_catalogo, st.session_state.stocks)
                                if resultados_ugreen and len(resultados_ugreen) > 0:
                                    prod = resultados_ugreen[0]
                                    precio = prod['precios'].get(st.session_state.precio_key, 0)
                                    if precio > 0 and prod['tiene_stock']:
                                        cantidad_cotizar = min(pedido['cantidad'], prod['stock'])
                                        estado = "✅ OK"
                                    elif precio > 0 and not prod['tiene_stock']:
                                        cantidad_cotizar = 0
                                        estado = "❌ Sin stock"
                                    else:
                                        cantidad_cotizar = 0
                                        estado = "❌ Sin precio"
                                    
                                    resultados_procesados.append({
                                        'sku': prod['sku'],
                                        'descripcion': prod['descripcion'],
                                        'precio': precio,
                                        'stock_total': prod['stock'],
                                        'tiene_stock': prod['tiene_stock'],
                                        'tiene_precio': precio > 0,
                                        'cantidad_solicitada': pedido['cantidad'],
                                        'cantidad_cotizar': cantidad_cotizar,
                                        'estado': estado,
                                        'tipo': 'UGREEN'
                                    })
                                else:
                                    resultados_procesados.append({
                                        'sku': pedido['sku'],
                                        'descripcion': f"SKU: {pedido['sku']}",
                                        'precio': 0,
                                        'stock_total': 0,
                                        'tiene_stock': False,
                                        'tiene_precio': False,
                                        'cantidad_solicitada': pedido['cantidad'],
                                        'cantidad_cotizar': 0,
                                        'estado': "❌ No encontrado",
                                        'tipo': 'UGREEN'
                                    })
                            
                            st.session_state.resultados_bulk = resultados_procesados
                            st.success(f"✅ Procesados {len(pedidos)} productos UGREEN")
    
    with col_b2:
        if st.button("📋 Agregar válidos al carrito", use_container_width=True):
            if hasattr(st.session_state, 'resultados_bulk') and st.session_state.resultados_bulk:
                agregados = 0
                for prod in st.session_state.resultados_bulk:
                    if prod['cantidad_cotizar'] > 0 and prod['tiene_precio']:
                        item_carrito = {
                            'sku': prod['sku'],
                            'descripcion': prod['descripcion'],
                            'cantidad': prod['cantidad_cotizar'],
                            'precio': prod['precio'],
                            'total': prod['precio'] * prod['cantidad_cotizar'],
                            'stock_yessica': prod.get('stock_yessica', 0),
                            'stock_apri004': prod.get('stock_apri004', 0),
                            'stock_apri001': prod.get('stock_apri001', 0),
                            'tipo': prod.get('tipo', 'XIAOMI')
                        }
                        st.session_state.carrito.append(item_carrito)
                        agregados += 1
                st.success(f"✅ Agregados {agregados} productos al carrito")
                st.rerun()
            else:
                st.warning("Primero procesa una lista")
    
    # Mostrar resultados bulk
    if 'resultados_bulk' in st.session_state and st.session_state.resultados_bulk:
        st.markdown("---")
        st.markdown("### 📋 Productos procesados")
        
        for prod in st.session_state.resultados_bulk:
            if prod.get('tipo') == 'UGREEN':
                badge_stock = f'<span class="badge-ugreen">📦 STOCK REAL: {prod["stock_total"]}</span>' if prod['stock_total'] > 0 else '<span class="badge-warning">❌ Sin stock</span>'
                
                st.markdown(f"""
                <div style="background:white;border-radius:16px;padding:1rem;margin-bottom:1rem;border-left:5px solid #00BCD4;box-shadow:0 2px 4px rgba(0,0,0,0.1);">
                    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;">
                        <div><strong style="color:#1a1a2e;">📦 SKU: {prod['sku']}</strong> <span style="background:#00BCD4;color:white;padding:2px 8px;border-radius:12px;">UGREEN</span></div>
                        <div><span style="background:#2196F3;color:white;padding:2px 8px;border-radius:12px;">Solicitado: {prod['cantidad_solicitada']}</span></div>
                    </div>
                    <div style="margin-top:8px;color:#1a1a2e;"><strong>📝 Descripción:</strong> {prod['descripcion'][:100]}</div>
                    <div style="margin-top:8px;color:#1a1a2e;">💰 Precio: <strong style="color:#e67e22;">S/ {prod['precio']:,.2f}</strong> | 📦 Stock: <strong>{prod['stock_total']}</strong></div>
                    <div style="margin-top:8px;">{badge_stock}</div>
                    <div style="margin-top:8px;color:#1a1a2e;"><strong>📌 Estado:</strong> {prod['estado']}</div>
                </div>
                """, unsafe_allow_html=True)
            else:
                badge_stock = construir_badge_stock(prod['stock_yessica'], prod['stock_apri004'], prod['stock_apri001'])
                
                if prod['tiene_stock'] and not prod['tiene_precio']:
                    st.markdown(f"""
                    <div style="background:white;border-radius:16px;padding:1rem;margin-bottom:1rem;border-left:5px solid #f44336;box-shadow:0 2px 4px rgba(0,0,0,0.1);">
                        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;">
                            <div><strong style="color:#1a1a2e;">📦 SKU: {prod['sku']}</strong> <span style="background:#f44336;color:white;padding:2px 8px;border-radius:12px;">⚠️ ERROR DE SKU</span></div>
                            <div><span style="background:#ff9800;color:white;padding:2px 8px;border-radius:12px;">Solicitado: {prod['cantidad_solicitada']}</span></div>
                        </div>
                        <div style="margin-top:8px;color:#1a1a2e;"><strong>📝 Descripción:</strong> {prod['descripcion']}</div>
                        <div style="margin-top:8px;color:#1a1a2e;"><strong>📦 Stock disponible:</strong></div>
                        <div>{badge_stock}</div>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    if prod.get('sku_equivalente'):
                        st.markdown(f"""
                        <div style="background:#E8F5E9;border-radius:12px;padding:1rem;margin:0.5rem 0;border-left:4px solid #4CAF50;">
                            <strong style="color:#2E7D32;">💡 SKU EQUIVALENTE SUGERIDO</strong><br>
                            <span style="color:#1a1a2e;"><strong>SKU:</strong> <code>{prod['sku_equivalente']}</code><br>
                            <strong>Precio:</strong> S/ {prod.get('precio_equivalente', 0):,.2f}<br>
                            <strong>Coincidencia:</strong> {prod.get('similitud_equivalente', 0):.0f}%</span>
                        </div>
                        """, unsafe_allow_html=True)
                
                elif prod['tiene_stock'] and prod['tiene_precio']:
                    st.markdown(f"""
                    <div style="background:white;border-radius:16px;padding:1rem;margin-bottom:1rem;border-left:5px solid #4CAF50;box-shadow:0 2px 4px rgba(0,0,0,0.1);">
                        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;">
                            <div><strong style="color:#1a1a2e;">📦 SKU: {prod['sku']}</strong> <span style="background:#4CAF50;color:white;padding:2px 8px;border-radius:12px;">✅ CON STOCK Y PRECIO</span></div>
                            <div><span style="background:#2196F3;color:white;padding:2px 8px;border-radius:12px;">Cotizar: {prod['cantidad_cotizar']}/{prod['cantidad_solicitada']}</span></div>
                        </div>
                        <div style="margin-top:8px;color:#1a1a2e;"><strong>📝 Descripción:</strong> {prod['descripcion']}</div>
                        <div style="margin-top:8px;color:#1a1a2e;">💰 Precio: <strong style="color:#e67e22;">S/ {prod['precio']:,.2f}</strong></div>
                        <div>{badge_stock}</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                elif not prod['tiene_stock'] and prod['tiene_precio']:
                    st.markdown(f"""
                    <div style="background:white;border-radius:16px;padding:1rem;margin-bottom:1rem;border-left:5px solid #2196F3;box-shadow:0 2px 4px rgba(0,0,0,0.1);">
                        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;">
                            <div><strong style="color:#1a1a2e;">📦 SKU: {prod['sku']}</strong> <span style="background:#2196F3;color:white;padding:2px 8px;border-radius:12px;">📋 SOLO PRECIO</span></div>
                        </div>
                        <div style="margin-top:8px;color:#1a1a2e;"><strong>📝 Descripción:</strong> {prod['descripcion']}</div>
                        <div style="margin-top:8px;color:#1a1a2e;">💰 Precio: <strong style="color:#e67e22;">S/ {prod['precio']:,.2f}</strong></div>
                        <div>{badge_stock}</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                else:
                    st.markdown(f"""
                    <div style="background:white;border-radius:16px;padding:1rem;margin-bottom:1rem;border-left:5px solid #9e9e9e;box-shadow:0 2px 4px rgba(0,0,0,0.1);">
                        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;">
                            <div><strong style="color:#1a1a2e;">📦 SKU: {prod['sku']}</strong> <span style="background:#9e9e9e;color:white;padding:2px 8px;border-radius:12px;">❌ NO DISPONIBLE</span></div>
                        </div>
                        <div style="margin-top:8px;color:#1a1a2e;"><strong>📝 Descripción:</strong> {prod['descripcion']}</div>
                        <div>{badge_stock}</div>
                    </div>
                    """, unsafe_allow_html=True)
        
        st.divider()
        
        if st.button("🗑️ Limpiar resultados", key="clear_bulk_results", use_container_width=True):
            del st.session_state.resultados_bulk
            st.rerun()

# ========== TAB 2: BÚSQUEDA INTELIGENTE ==========
with tab2:
    st.markdown("### 🔍 Buscar productos por SKU o descripción")
    st.caption(f"🔎 Modo: **{st.session_state.modo}** | Busca por SKU o cualquier palabra")
    
    busqueda = st.text_input("", placeholder="Ej: 'RN0200065BK8' o 'Type-C Earphones'")
    
    if busqueda and len(busqueda) >= 2:
        if st.session_state.modo == "XIAOMI" and st.session_state.catalogos and st.session_state.stocks:
            with st.spinner("🔍 Buscando..."):
                productos_por_sku = {}
                
                # Buscar en catálogos
                for cat in st.session_state.catalogos:
                    df = cat['df']
                    col_sku = cat['col_sku']
                    col_desc = cat.get('col_desc')
                    
                    mask_sku = df[col_sku].astype(str).str.contains(busqueda, case=False, na=False)
                    mask_desc = pd.Series([False] * len(df))
                    if col_desc:
                        mask_desc = df[col_desc].astype(str).str.contains(busqueda, case=False, na=False)
                    mask = mask_sku | mask_desc
                    
                    for _, row in df[mask].iterrows():
                        sku = str(row[col_sku]).strip().upper()
                        descripcion = str(row[col_desc])[:200] if col_desc else f"SKU: {sku}"
                        precio = 0.0
                        if st.session_state.precio_key in cat['precios']:
                            col_precio = cat['precios'][st.session_state.precio_key]
                            precio = corregir_numero(row[col_precio])
                        stock_info = buscar_stock_para_sku(sku, st.session_state.stocks)
                        
                        if sku in productos_por_sku:
                            existente = productos_por_sku[sku]
                            if precio > existente['precio']:
                                existente['precio'] = precio
                            # ASIGNAR, NO SUMAR
                            if stock_info['yessica'] > 0:
                                existente['stock_yessica'] = stock_info['yessica']
                            if stock_info['apri004'] > 0:
                                existente['stock_apri004'] = stock_info['apri004']
                            if stock_info['apri001'] > 0:
                                existente['stock_apri001'] = stock_info['apri001']
                            existente['stock_total'] = existente['stock_yessica'] + existente['stock_apri004'] + existente['stock_apri001']
                            existente['tiene_stock'] = existente['stock_total'] > 0
                        else:
                            productos_por_sku[sku] = {
                                'sku': sku,
                                'descripcion': descripcion,
                                'precio': precio,
                                'stock_yessica': stock_info['yessica'],
                                'stock_apri004': stock_info['apri004'],
                                'stock_apri001': stock_info['apri001'],
                                'stock_total': stock_info['total'],
                                'tiene_stock': stock_info['total'] > 0,
                                'tiene_precio': precio > 0
                            }
                
                if productos_por_sku:
                    st.success(f"✅ {len(productos_por_sku)} productos encontrados")
                    resultados_lista = list(productos_por_sku.values())
                    resultados_lista.sort(key=lambda x: (-x['tiene_stock'], -x['tiene_precio']))
                    
                    for prod in resultados_lista:
                        badge_stock = construir_badge_stock(prod['stock_yessica'], prod['stock_apri004'], prod['stock_apri001'])
                        
                        if prod['tiene_stock'] and prod['tiene_precio']:
                            color_borde = "#4CAF50"
                            estado = "✅ CON STOCK Y PRECIO"
                        elif prod['tiene_stock'] and not prod['tiene_precio']:
                            color_borde = "#f44336"
                            estado = "⚠️ STOCK SIN PRECIO"
                        elif not prod['tiene_stock'] and prod['tiene_precio']:
                            color_borde = "#2196F3"
                            estado = "📋 SOLO PRECIO"
                        else:
                            color_borde = "#9e9e9e"
                            estado = "❌ NO DISPONIBLE"
                        
                        st.markdown(f"""
                        <div style="background:white;border-radius:16px;padding:1rem;margin-bottom:1rem;border-left:5px solid {color_borde};box-shadow:0 2px 4px rgba(0,0,0,0.1);">
                            <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;">
                                <div><strong style="color:#1a1a2e;">📦 SKU: {prod['sku']}</strong></div>
                                <span style="background:{color_borde};color:white;padding:2px 8px;border-radius:12px;">{estado}</span>
                            </div>
                            <div style="margin-top:8px;color:#1a1a2e;"><strong>📝 Descripción:</strong> {prod['descripcion']}</div>
                            <div style="margin-top:8px;color:#1a1a2e;"><strong>💰 Precio {st.session_state.precio_key}:</strong> <strong style="color:#e67e22;">S/ {prod['precio']:,.2f}</strong></div>
                            <div>{badge_stock}</div>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        if prod['tiene_stock'] and prod['tiene_precio']:
                            col_cant, col_btn = st.columns([1, 2])
                            with col_cant:
                                cantidad = st.number_input("Cantidad", min_value=1, max_value=prod['stock_total'], value=1, step=1, key=f"search_{prod['sku']}")
                            with col_btn:
                                if st.button(f"➕ Agregar a cotización", key=f"add_search_{prod['sku']}"):
                                    item = {
                                        'sku': prod['sku'],
                                        'descripcion': prod['descripcion'],
                                        'cantidad': cantidad,
                                        'precio': prod['precio'],
                                        'total': prod['precio'] * cantidad,
                                        'stock_yessica': prod['stock_yessica'],
                                        'stock_apri004': prod['stock_apri004'],
                                        'stock_apri001': prod['stock_apri001']
                                    }
                                    st.session_state.carrito.append(item)
                                    st.success(f"✅ Agregado {cantidad}x {prod['sku']}")
                                    st.rerun()
                        
                        st.divider()
                else:
                    st.info("No se encontraron productos")
        
        elif st.session_state.modo == "UGREEN" and st.session_state.ugreen_catalogo:
            with st.spinner("🔍 Buscando en UGREEN..."):
                resultados_ugreen = buscar_ugreen_producto(busqueda, st.session_state.ugreen_catalogo, st.session_state.stocks)
                if resultados_ugreen:
                    st.success(f"✅ {len(resultados_ugreen)} productos encontrados")
                    for prod in resultados_ugreen:
                        precio = prod['precios'].get(st.session_state.precio_key, 0)
                        
                        # Diferente badge según si tiene precio o no
                        if prod.get('desde') == 'stock_solo':
                            # Producto encontrado SOLO en stock (sin precio en catálogo)
                            badge_stock = f'<span class="badge-ugreen">📦 STOCK REAL: {prod["stock"]}</span>' if prod['stock'] > 0 else '<span class="badge-warning">❌ Sin stock</span>'
                            estado_badge = '<span style="background:#FF9800;color:white;padding:2px 8px;border-radius:12px;">⚠️ SOLO STOCK (sin precio)</span>'
                        else:
                            # Producto normal con precio
                            badge_stock = f'<span class="badge-ugreen">📦 STOCK REAL: {prod["stock"]}</span>' if prod['stock'] > 0 else '<span class="badge-warning">❌ Sin stock</span>'
                            estado_badge = ''
                        
                        st.markdown(f"""
                        <div style="background:white;border-radius:16px;padding:1rem;margin-bottom:1rem;border-left:5px solid #00BCD4;">
                            <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;">
                                <div><strong style="color:#1a1a2e;">📦 SKU: {prod['sku']}</strong> <span style="background:#00BCD4;color:white;padding:2px 8px;border-radius:12px;">UGREEN</span> {estado_badge}</div>
                            </div>
                            <div style="margin-top:8px;color:#1a1a2e;"><strong>📝 Descripción:</strong> {prod['descripcion']}</div>
                            <div style="margin-top:8px;color:#1a1a2e;">💰 Precio: <strong style="color:#e67e22;">S/ {precio:,.2f}</strong></div>
                            <div style="margin-top:8px;">{badge_stock}</div>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        if prod['tiene_stock']:
                            # Permitir agregar aunque no tenga precio (con precio 0)
                            col_cant, col_btn = st.columns([1, 2])
                            with col_cant:
                                max_stock = prod['stock']
                                cantidad = st.number_input("Cantidad", min_value=1, max_value=max_stock, value=1, step=1, key=f"ugreen_{prod['sku']}")
                            with col_btn:
                                if st.button(f"➕ Agregar a cotización", key=f"add_ugreen_{prod['sku']}"):
                                    item = {
                                        'sku': prod['sku'],
                                        'descripcion': prod['descripcion'],
                                        'cantidad': cantidad,
                                        'precio': precio,
                                        'total': precio * cantidad,
                                        'tipo': 'UGREEN'
                                    }
                                    st.session_state.carrito.append(item)
                                    st.success(f"✅ Agregado {cantidad}x {prod['sku']}")
                                    st.rerun()
                        st.divider()
                else:
                    st.info("No se encontraron productos")

# ========== TAB 3: CARRITO ==========
with tab3:
    st.markdown("### 🛒 Cotización actual")
    
    if not st.session_state.carrito:
        st.info("No hay productos en el carrito")
    else:
        for idx, item in enumerate(st.session_state.carrito):
            col1, col2, col3, col4, col5, col6 = st.columns([2, 3, 1, 1, 1, 0.5])
            with col1:
                st.write(f"**{item['sku']}**")
            with col2:
                st.write(item['descripcion'][:50])
            with col3:
                nueva_cant = st.number_input("Cant", min_value=0, value=item['cantidad'], step=1, key=f"edit_{idx}", label_visibility="collapsed")
                if nueva_cant != item['cantidad']:
                    if nueva_cant == 0:
                        st.session_state.carrito.pop(idx)
                        st.rerun()
                    else:
                        item['cantidad'] = nueva_cant
                        item['total'] = item['precio'] * nueva_cant
            with col4:
                st.write(f"S/ {item['precio']:,.2f}")
            with col5:
                st.write(f"S/ {item['total']:,.2f}")
            with col6:
                if st.button("🗑️", key=f"del_{idx}"):
                    st.session_state.carrito.pop(idx)
                    st.rerun()
            
            if item.get('tipo') == 'UGREEN':
                st.markdown('<span class="badge-ugreen">📦 UGREEN</span>', unsafe_allow_html=True)
            else:
                badge = construir_badge_stock(item.get('stock_yessica', 0), item.get('stock_apri004', 0), item.get('stock_apri001', 0))
                st.markdown(f'<div style="margin-bottom:0.5rem;">{badge}</div>', unsafe_allow_html=True)
            st.divider()
        
        total_general = sum(item['total'] for item in st.session_state.carrito)
        st.markdown(f"""
        <div style="background:linear-gradient(135deg,#e94560 0%,#c73e54 100%);border-radius:12px;padding:1rem;margin:1rem 0;text-align:center;">
            <span style="color:white;font-size:1.5rem;font-weight:bold;">TOTAL: S/ {total_general:,.2f}</span>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("### 📋 Datos del cliente")
        col_cli1, col_cli2 = st.columns(2)
        with col_cli1:
            cliente = st.text_input("Nombre del cliente", placeholder="Ej: Empresa SAC")
        with col_cli2:
            ruc = st.text_input("RUC/DNI", placeholder="Ej: 20123456789")
        
        col_exp1, col_exp2, col_exp3 = st.columns(3)
        with col_exp1:
            if st.button("📥 Exportar Excel", type="primary", use_container_width=True):
                if cliente:
                    items_export = [{'sku': i['sku'], 'descripcion': i['descripcion'], 'cantidad': i['cantidad'], 'precio': i['precio'], 'total': i['total']} for i in st.session_state.carrito]
                    excel = generar_excel(items_export, cliente, ruc)
                    st.download_button("💾 Descargar", data=excel, file_name=f"Cotizacion_{cliente}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx", use_container_width=True)
                    st.balloons()
                    st.success("✅ Cotización generada")
                else:
                    st.warning("Ingresa el nombre del cliente")
        with col_exp2:
            if st.button("📋 Copiar CSV", use_container_width=True):
                csv_text = "SKU,Descripción,Cantidad,Precio,Subtotal\n"
                for item in st.session_state.carrito:
                    csv_text += f"{item['sku']},{item['descripcion']},{item['cantidad']},{item['precio']:.2f},{item['total']:.2f}\n"
                csv_text += f"TOTAL,{total_general:.2f}"
                st.code(csv_text, language="csv")
        with col_exp3:
            if st.button("🧹 Limpiar todo", use_container_width=True):
                st.session_state.carrito = []
                st.rerun()
# ============================================
# NUEVA PESTAÑA #4: BÚSQUEDA INTELIGENTE POR CÓDIGO (MODEL MARK / NO. / SKU)
# ============================================

with tab4:
    st.markdown("""
    <div style="background:linear-gradient(135deg, #667eea 0%, #764ba2 100%); border-radius:12px; padding:1rem; margin-bottom:1.5rem;">
        <h3 style="color:white; margin:0;">🔎 SMART CODE SEARCH</h3>
        <p style="color:rgba(255,255,255,0.9); margin:0;">Busca por MODEL MARK (ej: PB532), NO. (ej: 20956A) o cualquier código</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Input con sugerencias
    col_search, col_mode_info = st.columns([3, 1])
    with col_search:
        busqueda_codigo = st.text_input(
            "", 
            placeholder="🔍 Ingresa cualquier código: PB532 | 20956A | WS211 | CN9405860NA8 | HiTune S3",
            label_visibility="collapsed"
        )
    with col_mode_info:
        modo_actual = st.session_state.modo
        st.info(f"📌 Modo actual: **{modo_actual}**")
    
    # Botones de ejemplo rápido
    st.caption("💡 Ejemplos rápidos:")
    col_ej1, col_ej2, col_ej3, col_ej4, col_ej5 = st.columns(5)
    ejemplos = ["PB532", "20956A", "WS211", "HiTune S3", "CN9405860NA8"]
    for col, ej in zip([col_ej1, col_ej2, col_ej3, col_ej4, col_ej5], ejemplos):
        with col:
            if st.button(ej, key=f"ej_{ej}", use_container_width=True):
                busqueda_codigo = ej
                st.rerun()
    
    if busqueda_codigo and len(busqueda_codigo) >= 2:
        
        # ========== BÚSQUEDA PARA XIAOMI ==========
        if st.session_state.modo == "XIAOMI" and st.session_state.catalogos and st.session_state.stocks:
            with st.spinner(f"🔍 Buscando '{busqueda_codigo}' en SKU, MODEL MARK, NO. y descripciones..."):
                
                busqueda_limpia = busqueda_codigo.strip().upper()
                productos_encontrados = {}
                detalles_busqueda = {'sku': 0, 'model_mark': 0, 'no': 0, 'descripcion': 0}
                
                for cat in st.session_state.catalogos:
                    df = cat['df']
                    col_sku = cat['col_sku']
                    col_desc = cat.get('col_desc')
                    
                    # Detectar columnas
                    col_model = None
                    col_no = None
                    for col in df.columns:
                        col_upper = str(col).upper()
                        if 'MODEL MARK' in col_upper or col_upper == 'MODEL':
                            col_model = col
                        if col_upper == 'NO.' or col_upper == 'NO':
                            col_no = col
                    
                    # Máscaras por tipo de búsqueda
                    mask_total = pd.Series([False] * len(df))
                    
                    # 1. Búsqueda en SKU
                    mask_sku = df[col_sku].astype(str).str.contains(busqueda_limpia, case=False, na=False)
                    mask_total = mask_total | mask_sku
                    if mask_sku.any():
                        detalles_busqueda['sku'] += mask_sku.sum()
                    
                    # 2. Búsqueda en MODEL MARK
                    if col_model:
                        mask_model = df[col_model].astype(str).str.contains(busqueda_limpia, case=False, na=False)
                        mask_total = mask_total | mask_model
                        if mask_model.any():
                            detalles_busqueda['model_mark'] += mask_model.sum()
                    
                    # 3. Búsqueda en NO.
                    if col_no:
                        mask_no = df[col_no].astype(str).str.contains(busqueda_limpia, case=False, na=False)
                        mask_total = mask_total | mask_no
                        if mask_no.any():
                            detalles_busqueda['no'] += mask_no.sum()
                    
                    # 4. Búsqueda en descripción
                    if col_desc:
                        mask_desc = df[col_desc].astype(str).str.contains(busqueda_limpia, case=False, na=False)
                        mask_total = mask_total | mask_desc
                        if mask_desc.any():
                            detalles_busqueda['descripcion'] += mask_desc.sum()
                    
                    # Procesar resultados
                    for _, row in df[mask_total].iterrows():
                        sku = str(row[col_sku]).strip().upper()
                        
                        if sku in productos_encontrados:
                            continue
                        
                        model_mark = ""
                        if col_model and pd.notna(row[col_model]):
                            model_mark = str(row[col_model]).strip()
                        
                        no_code = ""
                        if col_no and pd.notna(row[col_no]):
                            no_code = str(row[col_no]).strip()
                        
                        descripcion = str(row[col_desc])[:200] if col_desc else f"SKU: {sku}"
                        
                        precio = 0.0
                        if st.session_state.precio_key in cat['precios']:
                            col_precio = cat['precios'][st.session_state.precio_key]
                            precio = corregir_numero(row[col_precio])
                        
                        stock_info = buscar_stock_para_sku(sku, st.session_state.stocks)
                        
                        # Determinar cómo se encontró
                        encontrado_por = []
                        if busqueda_limpia in sku:
                            encontrado_por.append("SKU")
                        if model_mark and busqueda_limpia in model_mark:
                            encontrado_por.append("MODEL MARK")
                        if no_code and busqueda_limpia in no_code:
                            encontrado_por.append("NO.")
                        if col_desc and busqueda_limpia in descripcion.upper():
                            encontrado_por.append("Descripción")
                        
                        productos_encontrados[sku] = {
                            'sku': sku,
                            'model_mark': model_mark,
                            'no_code': no_code,
                            'descripcion': descripcion,
                            'precio': precio,
                            'stock_yessica': stock_info['yessica'],
                            'stock_apri004': stock_info['apri004'],
                            'stock_apri001': stock_info['apri001'],
                            'stock_total': stock_info['total'],
                            'tiene_stock': stock_info['total'] > 0,
                            'tiene_precio': precio > 0,
                            'encontrado_por': encontrado_por
                        }
                
                # Mostrar resumen de búsqueda
                st.markdown(f"""
                <div style="background:rgba(102,126,234,0.1); border-radius:12px; padding:0.8rem; margin-bottom:1rem;">
                    <div style="display:flex; justify-content:space-between; flex-wrap:wrap;">
                        <span>🔍 Búsqueda: <strong>'{busqueda_codigo}'</strong></span>
                        <span>📊 Resultados: <strong>{len(productos_encontrados)} productos</strong></span>
                    </div>
                    <div style="display:flex; gap:1rem; margin-top:0.5rem; font-size:0.8rem;">
                        <span>📦 SKU: {detalles_busqueda['sku']}</span>
                        <span>🏷️ MODEL MARK: {detalles_busqueda['model_mark']}</span>
                        <span>🔢 NO.: {detalles_busqueda['no']}</span>
                        <span>📝 Descripción: {detalles_busqueda['descripcion']}</span>
                    </div>
                </div>
                """, unsafe_allow_html=True)
                
                if productos_encontrados:
                    # Ordenar: con stock y precio primero
                    resultados_lista = list(productos_encontrados.values())
                    resultados_lista.sort(key=lambda x: (-x['tiene_stock'], -x['tiene_precio']))
                    
                    for prod in resultados_lista:
                        # Determinar color según estado
                        if prod['tiene_stock'] and prod['tiene_precio']:
                            color_borde = "#4CAF50"
                            estado_texto = "✅ DISPONIBLE"
                            estado_color = "#4CAF50"
                        elif prod['tiene_stock'] and not prod['tiene_precio']:
                            color_borde = "#FF9800"
                            estado_texto = "⚠️ STOCK SIN PRECIO"
                            estado_color = "#FF9800"
                        elif not prod['tiene_stock'] and prod['tiene_precio']:
                            color_borde = "#2196F3"
                            estado_texto = "📋 SOLO PRECIO"
                            estado_color = "#2196F3"
                        else:
                            color_borde = "#9e9e9e"
                            estado_texto = "❌ NO DISPONIBLE"
                            estado_color = "#9e9e9e"
                        
                        # Badges de cómo se encontró
                        badges_encontrado = "".join([f'<span style="background:#667eea;color:white;padding:2px 6px;border-radius:12px;font-size:0.7rem;margin-right:4px;">{metodo}</span>' for metodo in prod['encontrado_por']])
                        
                        # Badges de stock
                        badge_stock = construir_badge_stock(prod['stock_yessica'], prod['stock_apri004'], prod['stock_apri001'])
                        
                        st.markdown(f"""
                        <div style="background:white;border-radius:16px;padding:1rem;margin-bottom:1rem;border-left:5px solid {color_borde};box-shadow:0 2px 8px rgba(0,0,0,0.1);">
                            <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;flex-wrap:wrap;">
                                <div>
                                    <strong style="color:#1a1a2e;font-size:1.1rem;">📦 {prod['sku']}</strong>
                                    <span style="background:{estado_color};color:white;padding:2px 10px;border-radius:20px;font-size:0.7rem;margin-left:8px;">{estado_texto}</span>
                                </div>
                                <div style="display:flex;gap:4px;">
                                    {badges_encontrado}
                                </div>
                            </div>
                            
                            <div style="display:grid; grid-template-columns:1fr 1fr; gap:8px; margin-bottom:10px;">
                                <div style="color:#1a1a2e; background:#f5f5f5; padding:4px 8px; border-radius:8px;">
                                    🏷️ <strong>MODEL MARK:</strong> {prod['model_mark'] or '-'}
                                </div>
                                <div style="color:#1a1a2e; background:#f5f5f5; padding:4px 8px; border-radius:8px;">
                                    🔢 <strong>NO.:</strong> {prod['no_code'] or '-'}
                                </div>
                            </div>
                            
                            <div style="margin-top:8px;color:#1a1a2e;">
                                <strong>📝 Descripción:</strong> {prod['descripcion']}
                            </div>
                            
                            <div style="margin-top:12px; display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap;">
                                <div>
                                    <span style="color:#1a1a2e;">💰 Precio <strong>{st.session_state.precio_key}</strong>:</span>
                                    <strong style="color:#e67e22; font-size:1.2rem; margin-left:8px;">S/ {prod['precio']:,.2f}</strong>
                                </div>
                                <div>
                                    {badge_stock}
                                </div>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        if prod['tiene_stock'] and prod['tiene_precio']:
                            col_cant, col_btn = st.columns([1, 3])
                            with col_cant:
                                cantidad = st.number_input(
                                    "Cantidad", 
                                    min_value=1, 
                                    max_value=prod['stock_total'], 
                                    value=min(1, prod['stock_total']), 
                                    step=1, 
                                    key=f"smart_{prod['sku']}"
                                )
                            with col_btn:
                                if st.button(f"➕ Agregar a cotización", key=f"add_smart_{prod['sku']}", use_container_width=True):
                                    item = {
                                        'sku': prod['sku'],
                                        'descripcion': prod['descripcion'],
                                        'cantidad': cantidad,
                                        'precio': prod['precio'],
                                        'total': prod['precio'] * cantidad,
                                        'stock_yessica': prod['stock_yessica'],
                                        'stock_apri004': prod['stock_apri004'],
                                        'stock_apri001': prod['stock_apri001']
                                    }
                                    st.session_state.carrito.append(item)
                                    st.success(f"✅ Agregado {cantidad}x {prod['sku']}")
                                    st.rerun()
                        st.markdown("---")
                else:
                    st.warning(f"❌ No se encontraron productos con '{busqueda_codigo}'")
        
        # ========== BÚSQUEDA PARA UGREEN ==========
        elif st.session_state.modo == "UGREEN" and st.session_state.ugreen_catalogo:
            with st.spinner(f"🔍 Buscando '{busqueda_codigo}' en UGREEN..."):
                
                df = st.session_state.ugreen_catalogo['df']
                col_sku = st.session_state.ugreen_catalogo['col_sku']
                col_desc = st.session_state.ugreen_catalogo['col_desc']
                
                # Detectar columnas
                col_model = None
                col_no = None
                for col in df.columns:
                    col_upper = str(col).upper()
                    if 'MODEL MARK' in col_upper or col_upper == 'MODEL':
                        col_model = col
                    if col_upper == 'NO.' or col_upper == 'NO':
                        col_no = col
                
                busqueda_limpia = busqueda_codigo.strip().upper()
                mask_total = pd.Series([False] * len(df))
                
                # Buscar en todas las columnas
                mask_total = mask_total | df[col_sku].astype(str).str.contains(busqueda_limpia, case=False, na=False)
                if col_model:
                    mask_total = mask_total | df[col_model].astype(str).str.contains(busqueda_limpia, case=False, na=False)
                if col_no:
                    mask_total = mask_total | df[col_no].astype(str).str.contains(busqueda_limpia, case=False, na=False)
                if col_desc:
                    mask_total = mask_total | df[col_desc].astype(str).str.contains(busqueda_limpia, case=False, na=False)
                
                productos_encontrados = []
                for _, row in df[mask_total].iterrows():
                    sku = str(row[col_sku]).strip().upper()
                    model_mark = ""
                    if col_model and pd.notna(row[col_model]):
                        model_mark = str(row[col_model]).strip()
                    no_code = ""
                    if col_no and pd.notna(row[col_no]):
                        no_code = str(row[col_no]).strip()
                    descripcion = str(row[col_desc])[:200] if col_desc else f"SKU: {sku}"
                    
                    precio_mayor = corregir_numero(row.get('Mayor', 0))
                    precio_caja = corregir_numero(row.get('Caja', 0))
                    precio_vip = corregir_numero(row.get('Vip', 0))
                    
                    stock = buscar_stock_ugreen(sku, st.session_state.stocks)
                    
                    productos_encontrados.append({
                        'sku': sku,
                        'model_mark': model_mark,
                        'no_code': no_code,
                        'descripcion': descripcion,
                        'precios': {'P. IR': precio_mayor, 'P. BOX': precio_caja, 'P. VIP': precio_vip},
                        'stock': stock,
                        'tiene_stock': stock > 0,
                        'tiene_precio': precio_vip > 0 or precio_caja > 0 or precio_mayor > 0
                    })
                
                if productos_encontrados:
                    st.success(f"✅ {len(productos_encontrados)} productos encontrados")
                    for prod in productos_encontrados:
                        precio = prod['precios'].get(st.session_state.precio_key, 0)
                        badge_stock = f'📦 STOCK REAL: {prod["stock"]}' if prod['stock'] > 0 else '❌ Sin stock'
                        
                        st.markdown(f"""
                        <div style="background:white;border-radius:16px;padding:1rem;margin-bottom:1rem;border-left:5px solid #00BCD4;">
                            <div><strong style="color:#1a1a2e;">📦 SKU: {prod['sku']}</strong> <span style="background:#00BCD4;color:white;padding:2px 8px;border-radius:12px;">UGREEN</span></div>
                            <div style="margin:8px 0;"><strong>🏷️ MODEL MARK:</strong> {prod['model_mark'] or '-'} | <strong>🔢 NO.:</strong> {prod['no_code'] or '-'}</div>
                            <div><strong>📝 Descripción:</strong> {prod['descripcion']}</div>
                            <div style="margin-top:8px;"><strong>💰 Precio {st.session_state.precio_key}:</strong> <strong style="color:#e67e22;">S/ {precio:,.2f}</strong></div>
                            <div style="margin-top:4px;">{badge_stock}</div>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        if prod['tiene_stock'] and prod['tiene_precio']:
                            col_cant, col_btn = st.columns([1, 3])
                            with col_cant:
                                cantidad = st.number_input("Cantidad", min_value=1, max_value=prod['stock'], value=1, key=f"ugreen_smart_{prod['sku']}")
                            with col_btn:
                                if st.button(f"➕ Agregar a cotización", key=f"add_ugreen_smart_{prod['sku']}"):
                                    st.session_state.carrito.append({
                                        'sku': prod['sku'],
                                        'descripcion': prod['descripcion'],
                                        'cantidad': cantidad,
                                        'precio': precio,
                                        'total': precio * cantidad,
                                        'tipo': 'UGREEN'
                                    })
                                    st.success(f"✅ Agregado {cantidad}x {prod['sku']}")
                                    st.rerun()
                        st.divider()
                else:
                    st.warning(f"❌ No se encontraron productos con '{busqueda_codigo}'")
        
        else:
            if st.session_state.modo == "XIAOMI":
                st.info("📂 Carga catálogos y stock primero")
            else:
                st.info("📂 Carga el catálogo UGREEN y stock primero")
    
    elif not busqueda_codigo:
        st.info("🔍 Ingresa un código para buscar (MODEL MARK, NO., SKU o descripción)")

# ============================================
# FOOTER
# ============================================

st.markdown("---")
st.markdown(f'<div class="footer">⚡ QTC Smart Sales Pro v4.5 | Modo: {st.session_state.modo} | YESSICA/APRI.004: stock inmediato | APRI.001: stock remoto | {datetime.now().strftime("%Y-%m-%d %H:%M")}</div>', unsafe_allow_html=True)
